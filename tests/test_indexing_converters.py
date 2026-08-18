"""Part 2: SourceDocumentConverter + classify→convert pipeline."""

from io import BytesIO

from docx import Document as DocxDocument
from haystack.dataclasses import ByteStream
from openpyxl import Workbook

from app.pipelines.indexing.document_converter import SourceDocumentConverter
from app.pipelines.indexing.mime_map import (
    MIME_CSV,
    MIME_DOCX,
    MIME_JSON,
    MIME_TEXT_PLAIN,
    MIME_XLSX,
    guess_mime_from_filename,
)
from app.pipelines.indexing.pipeline import build_indexing_pipeline, run_indexing_pipeline
from app.services.indexing import IndexingIngestService, byte_stream_from_upload


def _bs(name: str, data: bytes) -> ByteStream:
    return ByteStream(
        data=data,
        meta={"file_path": name, "filename": name},
        mime_type=guess_mime_from_filename(name),
    )


def test_convert_plain_text() -> None:
    out = SourceDocumentConverter().run(
        unstructured_sources=[_bs("brief.txt", b"Need a scissors lift")]
    )
    assert out["document_count"] == 1
    assert out["unstructured_document_count"] == 1
    assert "scissors" in (out["documents"][0].content or "")
    assert out["documents"][0].meta.get("data_kind") == "unstructured"


def test_convert_markdown() -> None:
    out = SourceDocumentConverter().run(
        unstructured_sources=[_bs("notes.md", b"# Site\n\nNeed boom lift")]
    )
    assert out["document_count"] == 1
    content = out["documents"][0].content or ""
    assert "boom" in content.lower() or "Site" in content


def test_convert_csv_structured() -> None:
    out = SourceDocumentConverter().run(
        structured_sources=[_bs("needs.csv", b"type,qty\nexcavator,2\n")]
    )
    assert out["structured_document_count"] == 1
    assert out["unstructured_document_count"] == 0
    assert "excavator" in (out["documents"][0].content or "")
    assert out["documents"][0].meta.get("data_kind") == "structured"
    assert out["documents"][0].meta.get("mime_type") == MIME_CSV


def test_convert_json_as_text() -> None:
    out = SourceDocumentConverter().run(
        structured_sources=[_bs("needs.json", b'{"equipment":"fork lift","qty":1}')]
    )
    assert out["structured_document_count"] == 1
    assert "fork lift" in (out["documents"][0].content or "")
    assert out["documents"][0].meta.get("mime_type") == MIME_JSON


def test_convert_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "type"
    ws["B1"] = "qty"
    ws["A2"] = "Fork Lift"
    ws["B2"] = 1
    buf = BytesIO()
    wb.save(buf)
    out = SourceDocumentConverter().run(structured_sources=[_bs("fleet.xlsx", buf.getvalue())])
    assert out["structured_document_count"] >= 1
    joined = " ".join(d.content or "" for d in out["documents"])
    assert "Fork Lift" in joined
    assert out["documents"][0].meta.get("mime_type") == MIME_XLSX


def test_convert_docx() -> None:
    doc = DocxDocument()
    doc.add_paragraph("Need scissors lift for indoor mezzanine")
    buf = BytesIO()
    doc.save(buf)
    out = SourceDocumentConverter().run(unstructured_sources=[_bs("spec.docx", buf.getvalue())])
    assert out["unstructured_document_count"] == 1
    assert "scissors" in (out["documents"][0].content or "").lower()
    assert out["documents"][0].meta.get("mime_type") == MIME_DOCX


def test_pipeline_classify_then_convert_txt() -> None:
    from haystack.document_stores.in_memory import InMemoryDocumentStore

    from app.pipelines.indexing.embedder_factory import build_document_embedder

    pipe = build_indexing_pipeline(
        document_store=InMemoryDocumentStore(),
        embedder=build_document_embedder(mode="mock", dimension=8),
    )
    out = run_indexing_pipeline(
        pipe,
        sources=[_bs("project.txt", b"Indoor elevated work for scissors lift")],
    )
    assert out["data_kind"] == "unstructured"
    assert out["documents_written"] >= 1
    assert MIME_TEXT_PLAIN in out["mime_types_seen"]


def test_pipeline_classify_then_convert_csv() -> None:
    from haystack.document_stores.in_memory import InMemoryDocumentStore

    from app.pipelines.indexing.embedder_factory import build_document_embedder

    pipe = build_indexing_pipeline(
        document_store=InMemoryDocumentStore(),
        embedder=build_document_embedder(mode="mock", dimension=8),
    )
    out = run_indexing_pipeline(
        pipe,
        sources=[_bs("needs.csv", b"type,qty\nBoom Lift,1\n")],
    )
    assert out["data_kind"] == "structured"
    assert out["documents_written"] >= 1


def test_service_returns_lean_summary() -> None:
    service = IndexingIngestService()
    result = service.ingest_from_project_spec(
        user_id="u_conv",
        project_text="Need one excavator for trench work",
    )
    assert result.ingest_id.startswith("ing_")
    assert result.user_id == "u_conv"
    assert "excavator" in result.user_requirement_summary.lower()
    assert isinstance(result.warnings, list)


def test_service_csv_has_structured_docs() -> None:
    service = IndexingIngestService()
    src = byte_stream_from_upload(
        raw=b"equipment,qty\nBoom Lift,1\n",
        filename="fleet.csv",
    )
    result = service.ingest_from_project_spec(user_id="u_conv", file_sources=[src])
    assert result.user_id == "u_conv"
    assert "boom" in result.user_requirement_summary.lower()


def test_empty_sources_convert_zero() -> None:
    out = SourceDocumentConverter().run(structured_sources=[], unstructured_sources=[])
    assert out["document_count"] == 0
    assert out["documents"] == []
